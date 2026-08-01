from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

NOTATION = re.compile(r"\b(?:LD|CM)\s*[-#:]?\s*\d+\b", re.IGNORECASE)


def analyse_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheets = []
    matches = []
    for sheet in workbook.worksheets:
        values = 0
        fills: dict[str, int] = {}
        comments = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values += 1
                    text = str(cell.value)
                    for match in NOTATION.finditer(text):
                        matches.append(
                            {
                                "sheet": sheet.title,
                                "cell": cell.coordinate,
                                "notation": match.group(),
                            }
                        )
                fill = cell.fill
                colour = fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.theme
                if fill.fill_type and colour is not None:
                    key = f"{fill.fill_type}:{colour}"
                    fills[key] = fills.get(key, 0) + 1
                if cell.comment:
                    comments.append(
                        {
                            "cell": cell.coordinate,
                            "author": cell.comment.author,
                            "text": cell.comment.text,
                        }
                    )
        sheets.append(
            {
                "title": sheet.title,
                "dimensions": sheet.calculate_dimension(),
                "non_empty_values": values,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "fills": fills,
                "comments": comments,
            }
        )
    return {"source": str(path), "sheets": sheets, "ld_cm_matches": matches}


def main() -> None:
    parser = argparse.ArgumentParser(description="Diagnose workbook structure without importing it")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.workbook.with_suffix(".diagnostic.json")
    output.write_text(json.dumps(analyse_workbook(args.workbook), indent=2) + "\n")
    print(f"Diagnostic written: {output}")


if __name__ == "__main__":
    main()
