from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.commands.seed import seed_development_data
from app.commands.workbook_diagnostic import analyse_workbook
from app.config import get_settings
from app.models.audit import AuditLog
from app.services.system import backup_database, core_export


def test_important_changes_are_audited(session: Session) -> None:
    seed_development_data(session)
    audit = session.scalars(select(AuditLog).where(AuditLog.entity_type == "jobs")).all()
    assert audit
    assert any(item.action == "create" for item in audit)


def test_database_backup_and_core_export(tmp_path: Path, monkeypatch, session: Session) -> None:  # type: ignore[no-untyped-def]
    source = tmp_path / "source.db"
    with sqlite3.connect(source) as connection:
        connection.execute("CREATE TABLE proof (value TEXT)")
        connection.execute("INSERT INTO proof VALUES ('safe')")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{source}")
    get_settings.cache_clear()
    try:
        backup = backup_database(tmp_path / "backups")
    finally:
        get_settings.cache_clear()
    with sqlite3.connect(backup) as connection:
        assert connection.execute("SELECT value FROM proof").fetchone() == ("safe",)

    seed_development_data(session)
    exported = core_export(session)
    assert exported["jobs"]
    assert exported["equipment_assets"]


def test_workbook_diagnostic_reads_structure_and_notation(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DAILY"
    sheet["A1"] = "LD 16 / CM17"
    sheet["A1"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["A1"].comment = Comment("Planner note", "Tester")
    sheet.merge_cells("B2:C2")
    path = tmp_path / "sample.xlsx"
    workbook.save(path)

    result = analyse_workbook(path)
    assert len(result["ld_cm_matches"]) == 2
    assert result["sheets"][0]["comments"][0]["text"] == "Planner note"
    assert result["sheets"][0]["merged_ranges"] == ["B2:C2"]
